"""Lightweight smoke tests for the SCL-90 PLOS ONE reproducibility package.

Run without raw data to verify package structure. To also run a lightweight raw
data check, set SCL90_TEST_DATA=/path/to/Data_collegestudent.csv before pytest.
The raw-data test intentionally uses the quick route and does not serve as the
submitted full recomputation route.
"""
from __future__ import annotations
import os, subprocess, sys
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]


def test_required_static_and_si_files_exist():
    required = [
        ROOT / "code" / "run_all.py",
        ROOT / "code" / "run_plos_one_analysis.py",
        ROOT / "code" / "static" / "archived_lasso_l1_item_ranking_hunan_all90.csv",
        ROOT / "supporting_info" / "S1_Table.xlsx",
        ROOT / "supporting_info" / "S6_Table_data_quality_exclusion_sensitivity.csv",
        ROOT / "supporting_info" / "S7_Table_topk_triage_bootstrap_ci.csv",
        ROOT / "supporting_info" / "S8_Table_selective_full_form_escalation.csv",
        ROOT / "supporting_info" / "S11_Table_nonoverlap_and_severity_validation.csv",
        ROOT / "supporting_info" / "S12_Table_fully_nested_item_selection_cv_summary.csv",
        ROOT / "LICENSE",
        ROOT / "CITATION.cff",
    ]
    missing = [str(p.relative_to(ROOT)) for p in required if not p.exists()]
    assert not missing, f"Missing required package files: {missing}"


def test_all_supporting_tables_exist():
    for i in range(1, 24):
        matches = list((ROOT / "supporting_info").glob(f"S{i}_Table*"))
        assert matches, f"Missing S{i} Table"


def test_packaged_s1_workbook_opens():
    try:
        from openpyxl import load_workbook
    except ImportError:
        return
    wb = load_workbook(ROOT / "supporting_info" / "S1_Table.xlsx", read_only=True)
    assert wb.sheetnames == ["summary", "folds"]
    wb.close()


def test_no_obsolete_user_facing_terms_in_outputs():
    banned = [
        "Quality-gate-first",
        "response_pattern_audit_first",
        "Diagnostic subgroup only",
        "quality-gated selective escalation",
        "Domain-order-balanced domain-order-balanced",
    ]
    paths = list((ROOT / "supporting_info").glob("*")) + [
        ROOT / "README.md",
        ROOT / "MANIFEST.md",
        ROOT / "outputs" / "analysis_run_config.json",
    ]
    for path in paths:
        if path.suffix.lower() == ".xlsx" or not path.exists():
            continue
        text = path.read_text(encoding="utf-8", errors="ignore")
        for term in banned:
            assert term not in text, f"{term} found in {path}"


def test_figure3_no_obsolete_response_pattern_audit_text():
    svg = ROOT / "outputs" / "figures" / "figure3_selective_escalation_workflow.svg"
    if not svg.exists():
        return
    text = svg.read_text(encoding="utf-8", errors="ignore")
    assert "Response-pattern audit" not in text
    assert "pattern checks" not in text


def test_run_all_help_exposes_recompute_and_include_tree():
    result = subprocess.run([sys.executable, str(ROOT / "code" / "run_all.py"), "--help"], text=True, capture_output=True, check=True)
    assert "--recompute-random-baseline" in result.stdout
    assert "--include-tree" in result.stdout
    assert "--skip-tree" not in result.stdout


def test_optional_minimal_pipeline_if_raw_data_available(tmp_path):
    data = os.environ.get("SCL90_TEST_DATA")
    if not data:
        return
    out = tmp_path / "outputs_smoke"
    subprocess.check_call([
        sys.executable, str(ROOT / "code" / "run_all.py"),
        "--data", data,
        "--out", str(out),
        "--clean",
        "--quick",
    ])
    for rel in [
        "tables/table1_sample_characteristics_quality.csv",
        "tables/table3_performance_by_item_count_and_selection.csv",
        "tables/table11_selective_full_form_escalation.csv",
        "tables/tableS_order_assumed_domain_balanced_item_selection.csv",
        "tables/tableS_response_quality_restricted_selective_escalation.csv",
        "tables/tableS_retrospective_response_quality_audit.csv",
        "tables/tableS_first_stage_observable_quality_flags.csv",
        "figures/figure1_burden_performance_frontier.png",
        "figures/figure1_burden_performance_frontier.tif",
        "figures/figure3_selective_escalation_workflow.png",
        "supporting_info/S1_Table.xlsx",
    ]:
        assert (out / rel).exists(), rel
    try:
        from openpyxl import load_workbook
    except ImportError:
        return
    wb = load_workbook(out / "supporting_info" / "S1_Table.xlsx", read_only=True)
    assert wb.sheetnames == ["summary", "folds"]
    wb.close()


def test_no_figure_numbers_above_five_are_packaged():
    for folder in [ROOT / "outputs" / "figures", ROOT / "figure_files_for_upload"]:
        if not folder.exists():
            continue
        for path in folder.iterdir():
            stem = path.stem.lower()
            digits = "".join(ch for ch in stem if ch.isdigit())
            if digits:
                assert int(digits[0]) <= 5, f"Unexpected obsolete figure-numbered artifact: {path}"


def test_upload_tiffs_are_rgb_without_alpha():
    from PIL import Image

    for path in sorted((ROOT / "figure_files_for_upload").glob("Fig*.tif")):
        with Image.open(path) as image:
            assert image.mode in {"RGB", "L"}, f"Alpha channel found in {path}: {image.mode}"
            assert tuple(round(value) for value in image.info.get("dpi", ())) == (600, 600)
